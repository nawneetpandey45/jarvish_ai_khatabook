from sklearn.linear_model import LinearRegression
import pickle
from collections import defaultdict

import os
import sqlite3
import json
import speech_recognition as sr
import pyttsx3
from datetime import datetime
import numpy as np
import time
import pywhatkit
from difflib import get_close_matches
import re
from openpyxl import Workbook, load_workbook
import qrcode
import threading
def smart_match(i, l):
    if not i:
        return False
    try:
        return bool(get_close_matches(i, l, 1, 0.5)) or any(x in i for x in l)
    except:
        return False


# ---------- 1. CONFIG & DATABASE ----------
DB_FILE = "Jarvish_Master.db"
ADMIN_VOICE_CODE = "nawneet pandey"
EXCEL_FILE = "daily_sales.xlsx"
CLOUD_FILE = "cloud_backup.json"

item_prices = {
    "Maggi (packet)": 15,
    "Biscuit (packet)": 10,
    "Sugar (kg)": 45,
    "Bread (packet)": 40,
    "Rice (kg)": 50,
    "Dal (kg)": 120,
    "Oil (litre)": 150,
    "Salt (kg)": 20,
    "Tea (250gm)": 180,
    "Milk (litre)": 60,
    "Egg (piece)": 6,
    "Soap (piece)": 30,
    "Shampoo (sachet)": 5,
    "Toothpaste (tube)": 80,
    "Detergent (kg)": 90,
    "Potato (kg)": 25,
    "Onion (kg)": 30,
    "Tomato (kg)": 30,
    "Curd (500gm)": 35,
    "Butter (100gm)": 55,
    "Paneer (200gm)": 90
}


product_synonyms = {
    "cheeni":"Sugar","chini":"Sugar",
    "anda":"Egg","ande":"Egg",
    "doodh":"Milk","tel":"Oil","namak":"Salt"
}

stock = {k:100 for k in item_prices}

unit_map = {"kg":1,"kilo":1,"gram":0.001,"g":0.001,"ml":0.001,"liter":1,"packet":1}
hindi_nums = {"ek":"1","do":"2","teen":"3","char":"4","paanch":"5","aadha":"0.5","half":"0.5"}

def init_db():
    conn = sqlite3.connect(DB_FILE, check_same_thread=False)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS customers
        (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE, phone TEXT,
         address TEXT, face_encoding TEXT, quick_code TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS transactions
        (id INTEGER PRIMARY KEY AUTOINCREMENT, cust_id INTEGER, date TEXT,
         items TEXT, paid REAL, due REAL, FOREIGN KEY(cust_id) REFERENCES customers(id))""")
    conn.commit()
    return conn

db_conn = init_db()

def safe_db(q,p=()):
    try:
        cur=db_conn.cursor(); cur.execute(q,p); db_conn.commit(); return cur
    except Exception as e:
        try: db_conn.rollback()
        except: pass
        print("DB:",e); return None

# ---------- 2. HARDWARE & VOICE ----------
try:
    import cv2
    import face_recognition
    FACE_AVAILABLE = True
except:
    FACE_AVAILABLE = False

engine=pyttsx3.init(); engine.setProperty('rate',170)

def speak(t):
    print("Jarvish:",t)
    try: engine.say(t); engine.runAndWait()
    except: pass

def listen():
    r = sr.Recognizer()
    with sr.Microphone() as s:
        # Noise level adjust karein
        r.adjust_for_ambient_noise(s, duration=0.8)
        print("🎤 Sun raha hoon (Boliye)...")
        try:
            # Phrase time limit di hai taaki hamesha ke liye na atke
            audio = r.listen(s, timeout=5, phrase_time_limit=8)
            
            # Dono languages support karne ke liye en-IN behtar hai 
            # kyunki ye 'Maggi' ko English mein likhega jo aapke dict se match hoga
            text = r.recognize_google(audio, language="en-IN") 
            
            print("🗣️ You:", text)
            return text.lower()
        except sr.UnknownValueError:
            print("Jarvish: Awaz samajh nahi aayi")
            return ""
        except sr.RequestError:
            print("Jarvish: Internet check karein")
            return ""
        except Exception as e:
            return ""
    def smart_match(i, l):
     if not i:
        return False
    try:
        return bool(get_close_matches(i, l, 1, 0.5)) or any(x in i for x in l)
    except:
        return False


def get_input(p):
    print(f"\n👉 {p}")
    v=input(">> [Type OR Press Enter for Voice]: ").strip()
    return listen() if v=="" else v.lower()

# ---------- 3. FACE & CUSTOMER ----------
def register_face(name):
    if not FACE_AVAILABLE: return
    cap=cv2.VideoCapture(0); speak("Camera mein dekhein...")
    start=time.time()
    while time.time()-start<8:
        r,f=cap.read()
        if not r: continue
        s=f[::4,::4]
        rgb=np.ascontiguousarray(cv2.cvtColor(s,cv2.COLOR_BGR2RGB))
        encs=face_recognition.face_encodings(rgb)
        if encs:
            safe_db("UPDATE customers SET face_encoding=? WHERE name=?",(json.dumps(encs[0].tolist()),name))
            speak("Face link success!"); break
        cv2.imshow("Registering Face...",f)
        if cv2.waitKey(1)==ord('q'): break
    cap.release(); cv2.destroyAllWindows()

def get_customer(i):
    cur=safe_db("SELECT * FROM customers WHERE name=? OR quick_code=? OR phone=?",(i,i,i))
    r=cur.fetchone() if cur else None
    if not r:
        cur=safe_db("SELECT * FROM customers WHERE name LIKE ?",(f"%{i}%",))
        r=cur.fetchone() if cur else None
    return r

def login_by_face():
    if not FACE_AVAILABLE: return None
    cur=safe_db("SELECT id,name,face_encoding FROM customers WHERE face_encoding IS NOT NULL")
    u=cur.fetchall() if cur else []
    if not u: return None
    enc=[np.array(json.loads(x[2])) for x in u]; names=[x[1] for x in u]
    cap=cv2.VideoCapture(0); start=time.time(); found=None
    while time.time()-start<6:
        r,f=cap.read()
        if not r: continue
        s=f[::4,::4]; rgb=np.ascontiguousarray(cv2.cvtColor(s,cv2.COLOR_BGR2RGB))
        for e in face_recognition.face_encodings(rgb):
            m=face_recognition.compare_faces(enc,e,0.6)
            if True in m:
                found=get_customer(names[m.index(True)]); break
        if found: break
        cv2.imshow("Scanning Face...",f)
        if cv2.waitKey(1)==ord('q'): break
    cap.release(); cv2.destroyAllWindows()
    return found

def voice_face_identify():
    speak("Apna naam boliye")
    n=listen().strip()
    if not n: return None
    cur=safe_db("SELECT * FROM customers WHERE name LIKE ?",(f"%{n}%",))
    r=cur.fetchall() if cur else []
    if len(r)==1: return r[0]
    if len(r)>1: return login_by_face()
    return None

# ---------- SMART NLP MULTI-ITEM BILL ----------
def ai_sentence_bill(sentence):
    s = sentence.lower()

    for k, v in hindi_nums.items():
        s = s.replace(k, v)

    for k, v in product_synonyms.items():
        s = s.replace(k, v.lower())

    for w in ["aur", "and", "dedo", "lelo", "rupaye", "diya", "de", "do"]:
        s = s.replace(w, " ")

    s = s.replace("मैगी", "maggi").replace("बागी", "maggi")

    tokens = s.split()
    cart = []

    for i, t in enumerate(tokens):

        # 🔹 FIX 1: number tokens skip
        if t.replace('.', '', 1).isdigit():
            continue

        found_item = None
        for name in item_prices:
            if name.lower().startswith(t.lower()):
                found_item = name
                break

        if found_item:
            q = 1.0
            unit = 1.0

            # quantity
            if i-1 >= 0 and tokens[i-1].replace('.', '', 1).isdigit():
                q = float(tokens[i-1])

            # unit
            if i+1 < len(tokens) and tokens[i+1] in unit_map:
                unit = unit_map[tokens[i+1]]
            elif i+2 < len(tokens) and tokens[i+2] in unit_map:
                unit = unit_map[tokens[i+2]]

            cart.append((found_item, q * unit))

            # 🔹 FIX 2: item add hone ke baad loop roko
            

    # paid amount (simple version)
    nums = re.findall(r"[-+]?\d*\.\d+|\d+", s)
    paid = float(nums[-1]) if nums else 0

    return (cart, paid) if cart else None


# ---------- WHATSAPP + BULK ----------
def send_whatsapp(name,phone,due):
    if not phone or len(phone)<10: return False
    target=f"+91{phone}" if not phone.startswith('+') else phone
    try:
        pywhatkit.sendwhatmsg_instantly(target,f"Namaste {name}, aapka store par ₹{due} ka udhaar baki hai.",20,True,4)
        return True
    except: return False

def send_to_all_defaulters():
    cur=safe_db("""SELECT c.name,c.phone,SUM(t.due)
                   FROM customers c JOIN transactions t ON c.id=t.cust_id
                   GROUP BY c.id HAVING SUM(t.due)>0""")
    r=cur.fetchall() if cur else []
    if not r: speak("Kisi ka udhaar baki nahi."); return
    for n,p,d in r: send_whatsapp(n,p,d); time.sleep(2)
    speak("Sabko reminder chala gaya")

# ---------- STOCK + EXCEL ----------
def reduce_stock(cart):
    for i,q in cart:
        if i in stock:
            stock[i]-=q
            if stock[i]<0: stock[i]=0

def export_excel(d,c,s,p,du):
    try:
        if not os.path.exists(EXCEL_FILE):
            wb=Workbook(); ws=wb.active; ws.append(["Date","Customer","Items","Paid","Due"]); wb.save(EXCEL_FILE)
        wb=load_workbook(EXCEL_FILE); ws=wb.active
        ws.append([d,c,s,p,du]); wb.save(EXCEL_FILE)
    except Exception as e: print("Excel:",e)

# ---------- CLOUD ----------
def cloud_backup():
    try:
        cur=safe_db("SELECT * FROM transactions"); r=cur.fetchall() if cur else []
        with open(CLOUD_FILE,"w") as f: json.dump(r,f,indent=2)
    except Exception as e: print("Cloud:",e)

def auto_cloud():
    while True: cloud_backup(); time.sleep(120)

threading.Thread(target=auto_cloud,daemon=True).start()
# ================= NEW ADVANCED AI MODULES (ADD ONLY) =================

customer_memory = defaultdict(list)

def update_memory(cust_id, cart):
    for item,qty in cart:
        customer_memory[cust_id].append((item, qty, datetime.now().isoformat()))

def buying_pattern_ai(cust_id):
    cur = safe_db("SELECT items FROM transactions WHERE cust_id=?", (cust_id,))
    rows = cur.fetchall() if cur else []
    freq={}
    for r in rows:
        for it in r[0].split(","):
            k=it.strip().split(" ")[0]
            freq[k]=freq.get(k,0)+1
    if freq:
        speak(f"Ye customer mostly {max(freq,key=freq.get)} kharidta hai")

# ----- REAL ML DEMAND PREDICTION -----
ml_model = LinearRegression()

def train_demand_model():
    cur = safe_db("SELECT items FROM transactions")
    rows = cur.fetchall() if cur else []
    X=[]; y=[]
    for i,r in enumerate(rows):
        X.append([i])
        y.append(len(r[0].split(",")))
    if len(X)>5:
        ml_model.fit(X,y)
        pickle.dump(ml_model,open("demand.pkl","wb"))

def predict_demand():
    try:
        m=pickle.load(open("demand.pkl","rb"))
        p=m.predict([[100]])[0]
        speak(f"Kal approx {int(p)} items bik sakte hain")
    except:
        speak("Demand model ready nahi")

# ----- SIMPLE EMOTION DETECTION -----
def detect_emotion():
    if not FACE_AVAILABLE: return
    cap=cv2.VideoCapture(0)
    r,f=cap.read()
    cap.release()
    if not r: return
    g=cv2.cvtColor(f,cv2.COLOR_BGR2GRAY)
    speak("Customer happy lag raha hai" if g.mean()>120 else "Customer serious lag raha hai")

# ----- OFFLINE SYNC -----
OFFLINE_FILE="offline_queue.json"

def offline_save(tx):
    try:
        d=json.load(open(OFFLINE_FILE)) if os.path.exists(OFFLINE_FILE) else []
        d.append(tx)
        json.dump(d,open(OFFLINE_FILE,"w"))
    except: pass

def sync_offline():
    if not os.path.exists(OFFLINE_FILE): return
    for t in json.load(open(OFFLINE_FILE)):
        safe_db("INSERT INTO transactions (cust_id,date,items,paid,due) VALUES (?,?,?,?,?)",tuple(t))
    os.remove(OFFLINE_FILE)


# ---------- QR ----------
def generate_customer_qr(n,p):
    qrcode.make(f"{n}|{p}").save(f"{n}_qr.png")

# ---------- ORIGINAL SMART AI (kept, overridden below) ----------
def smart_suggestions():
    cur=safe_db("SELECT items FROM transactions ORDER BY id DESC LIMIT 20")
    r=cur.fetchall() if cur else []
    f={}
    for x in r:
        for it in x[0].split(","):
            k=it.strip().split(" ")[0]; f[k]=f.get(k,0)+1
    if f: speak("Aaj zyada bikne wala item lagta hai "+max(f,key=f.get))

# ---------- DASHBOARD ----------
def dashboard():
    cur=safe_db("SELECT COUNT(*),SUM(paid),SUM(due) FROM transactions")
    c,paid,due=cur.fetchone()
    print("\n📊 DASHBOARD"); print("Bills:",c); print("Cash:",paid or 0); print("Due:",due or 0); print("Stock:",stock)
    

# ---------- OVERRIDE SMART AI (PLACED BEFORE profile_menu) ----------
def smart_suggestions():
    cur = safe_db("SELECT items FROM transactions WHERE items!='CLEARANCE' ORDER BY id DESC LIMIT 30")
    rows = cur.fetchall() if cur else []
    freq = {}
    for r in rows:
        if not r or not r[0]: continue
        for it in r[0].split(","):
            k = it.strip().split(" ")[0]
            if k.upper()=="CLEARANCE": continue
            if k: freq[k]=freq.get(k,0)+1
    if freq: speak("Aaj zyada bikne wala item lagta hai "+max(freq,key=freq.get))
    else: speak("Abhi AI suggestion ke liye data kam hai")

# ---------- 5. PROFILE MENU ----------
def profile_menu(cust):
    cid,cname,cphone,_,_,_=cust
    while True:
        cur=safe_db("SELECT SUM(CASE WHEN due>0 THEN due ELSE 0 END) FROM transactions WHERE cust_id=?",(cid,))
        due=cur.fetchone()[0] if cur else 0; due=due or 0.0
        print(f"\n--- {cname.upper()} | KHATA: ₹{due} | No: {cphone} ---")
        print("1. Add Bill | 2. Clear | 3. Face Reg | 4. WhatsApp | 5. History | 6. QR | 7. AI | 8. Back")
        cmd=get_input("Action")

        if smart_match(cmd,["1","add","bill","बिल"]):
            txt=get_input("Bolo: 2 maggi aur aadha kilo sugar 30 diya")
            data=ai_sentence_bill(txt)
            if not data: speak("Samajh nahi aaya."); continue
            cart,paid=data; total=0; summ=[]
            for it,q in cart: total+=item_prices.get(it,15)*q; summ.append(f"{it} x{q}")
            speak(f"Total {total} rupaye. Confirm?")
            if "nahi" in listen(): speak("Cancel kar diya."); continue
            reduce_stock(cart)
            curr_due=total-paid
            if curr_due<0: speak(f"{abs(curr_due)} rupaye advance aaye")
            today=datetime.now().strftime("%d-%m-%Y")
            safe_db("INSERT INTO transactions (cust_id,date,items,paid,due) VALUES (?,?,?,?,?)",(cid,today,", ".join(summ),paid,curr_due))
            export_excel(today,cname,", ".join(summ),paid,curr_due); speak(f"Saved. Due: {curr_due}")
            update_memory(cid, cart)
            buying_pattern_ai(cid)
            detect_emotion()
            train_demand_model()


        elif smart_match(cmd,["4","whatsapp","भेज"]): send_whatsapp(cname,cphone,due)
        elif smart_match(cmd,["3","face","फोटो"]): register_face(cname)
        elif smart_match(cmd,["2","clear","saaf"]):
            safe_db("DELETE FROM transactions WHERE cust_id=?",(cid,))
            safe_db("INSERT INTO transactions (cust_id,date,items,paid,due) VALUES (?,?,?,?,?)",(cid,datetime.now().strftime("%d-%m-%Y"),"CLEARANCE",due,0))
            speak("Hisaab barabar. Khata reset ho gaya.")
        elif smart_match(cmd,["5","history"]):
            cur=safe_db("SELECT date,items,paid,due FROM transactions WHERE cust_id=? ORDER BY id DESC LIMIT 5",(cid,))
            for r in (cur.fetchall() if cur else []): print(f"{r[0]} | {r[1]} | Paid:{r[2]} | Due:{r[3]}")
        elif smart_match(cmd,["6","qr"]): generate_customer_qr(cname,cphone); speak("QR card ready")
        elif smart_match(cmd,["7","ai"]): smart_suggestions()
        elif smart_match(cmd,["8","back","वापस"]): break

# ---------- 6. MAIN LOOP ----------
def main_loop():
    speak("Jarvish Online. Password?")
    attempts=0
    while attempts<3:
        pwd=get_input("Password")
        if smart_match(pwd,[ADMIN_VOICE_CODE,"जार्विस कमान अल्फा"]): speak("Verified. Swagat hai boss!"); break
        attempts+=1; speak(f"Galat. {3-attempts} baki.")
    if attempts>=3: return
    while True:
        print("\n=== MASTER MENU ===")
        print("1. Profile | 2. New Cust | 3. Business Health | 4. Remind All | 5. Dashboard | 6. Exit")
        cmd=get_input("Command")
        if smart_match(cmd,["1","profile","प्रोफाइल"]):
            speak("Face scan ho raha hai...")
            cust=login_by_face() or voice_face_identify() or get_customer(get_input("Name/Code"))
            if cust: profile_menu(cust)
            else: speak("Customer nahi mila.")
        elif smart_match(cmd,["2","new","नया"]):
            name=get_input("Name?").capitalize(); phone=get_input("Phone?")
            safe_db("INSERT OR IGNORE INTO customers (name,phone,quick_code) VALUES (?,?,?)",(name,phone,phone[-4:] if phone else "0000"))
            speak(f"{name} register ho gaya.")
        elif smart_match(cmd,["3","health","summary"]):
            cur=safe_db("SELECT SUM(paid),SUM(due) FROM transactions"); res=cur.fetchone() if cur else (0,0)
            speak(f"Cash: {res[0] or 0}, Udhaar: {res[1] or 0}")
        elif smart_match(cmd,["4","remind","sabko"]): send_to_all_defaulters()

        elif smart_match(cmd,["5","dashboard"]): dashboard()

        elif smart_match(cmd,["predict"]):
         predict_demand()

        elif smart_match(cmd,["6","exit","band"]): break




if __name__=="__main__":
    main_loop()



# python "C:\Users\ASUS\Desktop\jarvish ai khatabook\khatabook.py"